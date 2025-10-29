import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import yaml

# ============================================================
# ⚙️ CONFIGURACIÓN DE RUTAS
# ============================================================
BASE_DIR = Path(__file__).resolve().parents[1]  # /UCC-SUPERVISION
CONFIG_PATH = BASE_DIR / "config/settings_anexo2.yaml"
CARPETA_RAW = BASE_DIR / "data/raw"

# ============================================================
# 📄 CARGAR CONFIGURACIÓN YAML
# ============================================================
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    CONFIG = yaml.safe_load(f)

# ============================================================
# 🧩 FUNCIONES AUXILIARES Y LOG
# ============================================================
def log(mensaje, tipo="INFO", anexo="ANEXO_2"):
    """Imprime un mensaje estándar con fecha, tipo y anexo."""
    hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    simbolo = {"INFO": "ℹ️", "OK": "✅", "WARN": "⚠️", "ERROR": "❌"}.get(tipo, "")
    print(f"{simbolo} [{anexo}] [{hora}] {mensaje}")

def extraer_valor(hoja, celda):
    valor = hoja[celda].value
    if isinstance(valor, str) and ":" in valor:
        return valor.split(":", 1)[1].strip()
    return valor

def limpiar_texto(v):
    if isinstance(v, str):
        return v.strip().title()
    return v

def formatear_fecha(v):
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%Y-%m-%d")
    return v

# ============================================================
# 📋 COLUMNAS COMUNES ESTANDARIZADAS
# ============================================================
COLUMNAS_COMUNES = [
    "Año", "Mes", "Región", "Archivo",
    "Unidad Territorial", "Distrito", "Supervisor", "Fecha Supervisión"
]

# ============================================================
# 🧩 PROCESAR UNA SOLA FICHA ANEXO 2
# ============================================================
def procesar_ficha(ruta_excel, region, mes, anio):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    hoja = wb.active

    meta = CONFIG["metadatos"]
    ut = limpiar_texto(extraer_valor(hoja, meta["unidad_territorial"])) or ""
    distrito = limpiar_texto(extraer_valor(hoja, meta["distrito"])) or ""
    supervisor = limpiar_texto(extraer_valor(hoja, meta["supervisor"])) or ""
    fecha = formatear_fecha(extraer_valor(hoja, meta["fecha"])) or ""

    # --- Ítems ---
    items_cfg = CONFIG["items"]
    items_dict = {}
    for rango in items_cfg["rangos"]:
        for fila in hoja.iter_rows(
            min_row=rango["fila_inicio"],
            max_row=rango["fila_fin"],
            min_col=items_cfg["col_inicio"],
            max_col=items_cfg["col_fin"],
            values_only=True
        ):
            if not fila or all(v is None for v in fila):
                continue
            pregunta = str(fila[0]).strip() if fila[0] else None
            calificaciones = fila[1:]
            if pregunta:
                valor = next((v for v in calificaciones if v not in [None, ""]), "NA")
                items_dict[pregunta[:120]] = valor

    if not items_dict:
        raise ValueError(f"La ficha {ruta_excel.name} no contiene ítems válidos.")

    # --- Cálculos ---
    valores_numericos = [v for v in items_dict.values() if isinstance(v, (int, float))]
    suma_total = sum(valores_numericos)
    total_items = len(valores_numericos)
    num_na = sum(1 for v in items_dict.values() if v == "NA")
    num_validos = len(items_dict) - num_na

    max_por_item = CONFIG["limites"]["max_por_item"]
    puntaje = round((suma_total / (total_items * max_por_item)) * 100, 1) if total_items > 0 else 0

    categoria = None
    for regla in sorted(CONFIG["clasificacion"], key=lambda x: x["max"]):
        if suma_total <= regla["max"]:
            categoria = regla["categoria"]
            break

    # --- Consolidar datos ---
    data = {
        "Año": anio,
        "Mes": mes,
        "Región": region,
        "Archivo": ruta_excel.name,
        "Unidad Territorial": ut,
        "Distrito": distrito,
        "Supervisor": supervisor,
        "Fecha Supervisión": fecha,
        **{f"Item_{i+1}": v for i, v in enumerate(items_dict.values())},
        "Ítems válidos": num_validos,
        "Ítems NA": num_na,
        "Suma Total": suma_total,
        "Puntaje (%)": puntaje,
        "Evaluación": categoria
    }

    return pd.DataFrame([data])

# ============================================================
# 🚀 PROCESAR TODAS LAS FICHAS DE TODOS LOS MESES Y REGIONES
# ============================================================
if not CARPETA_RAW.exists():
    raise FileNotFoundError(f"❌ No existe la carpeta: {CARPETA_RAW}")

registros = []
for carpeta_anio in sorted(CARPETA_RAW.iterdir()):
    if not carpeta_anio.is_dir():
        continue
    anio = carpeta_anio.name
    for carpeta_mes in sorted(carpeta_anio.iterdir()):
        if not carpeta_mes.is_dir():
            continue
        mes = carpeta_mes.name
        log(f"Procesando: {anio}/{mes}", "INFO")
        for carpeta_region in sorted(carpeta_mes.iterdir()):
            if not carpeta_region.is_dir():
                continue
            region = carpeta_region.name.upper()
            archivos = list(carpeta_region.glob("*ANEXO_2*.xlsx"))
            if not archivos:
                log(f"{anio}/{mes}/{region}: sin archivos Excel", "WARN")
                continue
            for archivo in archivos:
                if archivo.name.startswith("~$"):
                    continue
                try:
                    df_ficha = procesar_ficha(archivo, region, mes, anio)
                    registros.append(df_ficha)
                    log(f"{anio}/{mes}/{region}/{archivo.name} procesado", "OK")
                except Exception as e:
                    log(f"Error en {anio}/{mes}/{region}/{archivo.name}: {e}", "ERROR")

# ============================================================
# 💾 EXPORTAR CONSOLIDADO ACUMULATIVO
# ============================================================
if registros:
    df_total = pd.concat(registros, ignore_index=True)

    cols_final = COLUMNAS_COMUNES + [c for c in df_total.columns if c not in COLUMNAS_COMUNES]
    df_total = df_total[cols_final]

    salida_cfg = CONFIG["salida"]
    salida_dir = BASE_DIR / salida_cfg["carpeta"]
    salida_dir.mkdir(parents=True, exist_ok=True)

    nombre_archivo = "anexo2_consolidado.xlsx"
    salida_excel = salida_dir / nombre_archivo

    # Si ya existe un consolidado previo, acumula sin duplicar
    if salida_excel.exists():
        df_prev = pd.read_excel(salida_excel)
        df_total = pd.concat([df_prev, df_total], ignore_index=True)
        df_total.drop_duplicates(subset=["Archivo", "Región", "Mes", "Año"], inplace=True)

    df_total.to_excel(salida_excel, index=False)
    log(f"Consolidado actualizado: {salida_excel}", "OK")
    log(f"Total de fichas acumuladas: {len(df_total)}", "INFO")

else:
    log("No se procesó ninguna ficha.", "WARN")

log("Procesamiento de ANEXO 2 finalizado.", "OK")