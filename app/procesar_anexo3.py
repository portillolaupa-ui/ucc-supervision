import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import yaml

# ============================================================
# ⚙️ CONFIGURACIÓN DE RUTAS
# ============================================================
BASE_DIR = Path(__file__).resolve().parents[1]
CONFIG_PATH = BASE_DIR / "config/settings_anexo3.yaml"
CARPETA_RAW = BASE_DIR / "data/raw"

# ============================================================
# 📄 CARGAR CONFIGURACIÓN YAML
# ============================================================
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    CONFIG = yaml.safe_load(f)

# ============================================================
# 🧩 FUNCIONES AUXILIARES
# ============================================================
def log(mensaje, tipo="INFO", anexo="ANEXO_3"):
    """Imprime un log estandarizado con fecha, tipo y anexo."""
    hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    simbolo = {"INFO": "ℹ️", "OK": "✅", "WARN": "⚠️", "ERROR": "❌"}.get(tipo, "")
    print(f"{simbolo} [{anexo}] [{hora}] {mensaje}")

def extraer_valor(hoja, celda):
    valor = hoja[celda].value
    if isinstance(valor, str) and ":" in valor:
        return valor.split(":", 1)[1].strip()
    return valor

def limpiar_texto(v):
    return v.strip().title() if isinstance(v, str) else v

def formatear_fecha(v):
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%Y-%m-%d")
    return v

# ============================================================
# 📋 COLUMNAS COMUNES ESTANDARIZADAS
# ============================================================
COLUMNAS_COMUNES = [
    "Año", "Mes", "Región", "Archivo",
    "Responsable", "Facilitador", "Comunidad/Distrito",
    "Fecha Sesión", "Sesión Observada"
]

# ============================================================
# 🧩 PROCESAR UNA SOLA FICHA ANEXO 3
# ============================================================
def procesar_ficha(ruta_excel, region, mes, anio):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    hoja = wb.active

    meta = CONFIG["metadatos"]
    responsable = limpiar_texto(extraer_valor(hoja, meta["responsable"])) or ""
    comunidad = limpiar_texto(extraer_valor(hoja, meta["comunidad"])) or ""
    fecha_sesion = formatear_fecha(extraer_valor(hoja, meta["fecha_sesion"])) or ""
    sesion_obs = limpiar_texto(extraer_valor(hoja, meta["sesion_observada"])) or ""
    facilitador = limpiar_texto(extraer_valor(hoja, meta["facilitador"])) or ""

    registros = {s["nombre"]: [] for s in CONFIG["secciones"]}

    for s in CONFIG["secciones"]:
        nombre = s["nombre"]
        items = []

        for fila in hoja.iter_rows(
            min_row=s["fila_inicio"],
            max_row=s["fila_fin"],
            min_col=s["col_inicio"],
            max_col=s["col_fin"],
            values_only=True
        ):
            if not fila or all(v is None for v in fila):
                continue

            si = fila[-3] if len(fila) >= 3 else None
            parcial = fila[-2] if len(fila) >= 2 else None
            no = fila[-1] if len(fila) >= 1 else None

            if si is not None:
                valor = 2
            elif parcial is not None:
                valor = 1
            elif no is not None:
                valor = 0
            else:
                valor = "NA"

            items.append(valor)

        valores_numericos = [v for v in items if isinstance(v, (int, float))]
        suma_total = sum(valores_numericos)

        categoria = "Sin escala"
        for regla in sorted(s["clasificacion"], key=lambda x: x["max"]):
            if suma_total <= regla["max"]:
                categoria = regla["categoria"]
                break

        fila_data = {
            "Año": anio,
            "Mes": mes,
            "Región": region,
            "Archivo": ruta_excel.name,
            "Responsable": responsable,
            "Facilitador": facilitador,
            "Comunidad/Distrito": comunidad,
            "Fecha Sesión": fecha_sesion,
            "Sesión Observada": sesion_obs,
            **{f"Item_{i+1}": v for i, v in enumerate(items)},
            "Total": suma_total,
            "Evaluación": categoria
        }

        registros[nombre].append(fila_data)

    return registros

# ============================================================
# 🚀 PROCESAR TODAS LAS FICHAS
# ============================================================
consolidado_por_seccion = {s["nombre"]: [] for s in CONFIG["secciones"]}

if not CARPETA_RAW.exists():
    raise FileNotFoundError(f"❌ No existe la carpeta: {CARPETA_RAW}")

for carpeta_anio in sorted(CARPETA_RAW.iterdir()):
    if not carpeta_anio.is_dir():
        continue
    anio = carpeta_anio.name
    for carpeta_mes in sorted(carpeta_anio.iterdir()):
        if not carpeta_mes.is_dir():
            continue
        mes = carpeta_mes.name
        log(f"Procesando: {anio}/{mes}", "INFO")
        for carpeta_region in sorted(carpeta_mes.iterdir()):
            if not carpeta_region.is_dir():
                continue
            region = carpeta_region.name.upper()
            archivos = list(carpeta_region.glob("*ANEXO_3*.xlsx"))
            if not archivos:
                log(f"{anio}/{mes}/{region}: sin archivos Excel", "WARN")
                continue
            for archivo in archivos:
                if archivo.name.startswith("~$"):
                    continue
                try:
                    resultados = procesar_ficha(archivo, region, mes, anio)
                    for seccion, filas in resultados.items():
                        consolidado_por_seccion[seccion].extend(filas)
                    log(f"{anio}/{mes}/{region}/{archivo.name} procesado", "OK")
                except Exception as e:
                    log(f"Error procesando {anio}/{mes}/{region}/{archivo.name}: {e}", "ERROR")

# ============================================================
# 💾 EXPORTAR CONSOLIDADOS
# ============================================================
salida_cfg = CONFIG["salida"]
salida_dir = BASE_DIR / salida_cfg["carpeta"]
salida_dir.mkdir(parents=True, exist_ok=True)

for seccion, registros in consolidado_por_seccion.items():
    if not registros:
        continue

    df_nuevo = pd.DataFrame(registros)
    cols_final = COLUMNAS_COMUNES + [c for c in df_nuevo.columns if c not in COLUMNAS_COMUNES]
    df_nuevo = df_nuevo[cols_final]

    nombre_archivo = f"anexo3_{seccion.lower().replace(' ', '_')}_consolidado.xlsx"
    salida_excel = salida_dir / nombre_archivo

    if salida_excel.exists():
        df_prev = pd.read_excel(salida_excel)
        df_total = pd.concat([df_prev, df_nuevo], ignore_index=True)
        df_total.drop_duplicates(subset=["Archivo", "Región", "Mes", "Año"], inplace=True)
    else:
        df_total = df_nuevo

    df_total.to_excel(salida_excel, index=False)
    log(f"Consolidado actualizado: {salida_excel}", "OK")
    log(f"Total de registros en {seccion}: {len(df_total)}", "INFO")

log("Procesamiento de ANEXO 3 finalizado.", "OK")